#files/views.py
from rest_framework import viewsets, permissions, status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from .models import File, AuditLog, SystemSettings, EmployeeDirectory, DTRFile, DTREntry
from .serializers import FileSerializer, FileStatusSerializer, AuditLogSerializer, SystemSettingsSerializer, EmployeeDirectorySerializer, DTREntrySerializer, DTRFileSerializer
from accounts.permissions import ReadOnlyForViewer, IsOwnerOrAdmin, CanEditStatus, IsAdmin
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from django.http import FileResponse, Http404, HttpResponse
from rest_framework.permissions import IsAuthenticated, IsAdminUser
import csv
from openpyxl import load_workbook
import io
from accounts.models import User
from reportlab.pdfgen import canvas
from django.db.models import Count, Q
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from .utils import log_action, get_client_ip
from django.core.exceptions import ValidationError
from .utils import send_rejection_sms 
import pandas as pd
import math
from decimal import Decimal, InvalidOperation
from datetime import timedelta
import traceback

def log_action(user, action, status="success", ip_address=None):
    AuditLog.objects.create(
        user=user if user.is_authenticated else None,
        action=action,
        status=status,
        ip_address=ip_address
    )

class FileViewSet(viewsets.ModelViewSet):
    queryset = File.objects.all()
    serializer_class = FileSerializer
    permission_classes = [permissions.IsAuthenticated, CanEditStatus]

    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_parser_classes(self):
        if self.action == "create":
            return [MultiPartParser, FormParser]
        return [JSONParser]

    def get_queryset(self):
        user = self.request.user
        if user.role == "client":
            return File.objects.filter(owner=user).order_by("-uploaded_at")
        return File.objects.all().order_by("-uploaded_at")

    def get_permissions(self):
        if self.action == "destroy":
            permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
        elif self.action == "update_status":
            permission_classes = [IsAuthenticated, CanEditStatus]
        else:
            permission_classes = [IsAuthenticated]
        return [p() for p in permission_classes]
    
    def perform_create(self, serializer):
        user = self.request.user
        settings = SystemSettings.objects.first() 
        
        file_obj = serializer.validated_data['file']
        
        # 1. Check file size
        max_size = settings.max_file_size * 1024 * 1024 
        if file_obj.size > max_size:
            raise ValidationError(f"File exceeds max size of {settings.max_file_size} MB")
        
        # 2. Check allowed types
        ext = file_obj.name.split('.')[-1].lower()
        if ext not in settings.allowed_types:
            raise ValidationError(f"File type {ext} not allowed. Allowed: {settings.allowed_types}")
        
        # 3. Existing overwrite logic
        if user.role == 'client':
            existing_file = File.objects.filter(owner=user, file=file_obj.name).first()
            if existing_file:
                existing_file.file.delete(save=False)
                serializer.instance = existing_file
                serializer.save()
                log_action(user, f"updated file {file_obj.name}", ip_address=get_client_ip(self.request))
                return

        new_file = serializer.save(owner=user)
        log_action(user, f"uploaded file {new_file.file.name}", ip_address=get_client_ip(self.request))

    @action(
            detail=True, 
            methods=["get"], 
            permission_classes=[IsAuthenticated, ReadOnlyForViewer, IsOwnerOrAdmin]
    )

    @action(detail=True, methods=["get"], url_path="download")
    def download(self, request, pk=None):
        file = self.get_object()
        settings = SystemSettings.objects.first()
        
        if settings.require_verification and file.status != "verified":
            return Response({"detail": "File must be verified before download"}, status=403)
        
        if settings.log_downloads:
            log_action(request.user, f"downloaded file {file.file.name}", ip_address=get_client_ip(request))
        
        try:
            return FileResponse(file.file.open(), as_attachment=True, filename=file.file.name)
        except FileNotFoundError:
            raise Http404

    def perform_destroy(self, instance):
        settings = SystemSettings.objects.first()
        
        if settings.auto_archive:
            instance.status = "archived"
            instance.save(update_fields=["status"])
            log_action(self.request.user, f"archived file {instance.file.name}", ip_address=get_client_ip(self.request))
        else:
            file_name = instance.file.name
            super().perform_destroy(instance)
            log_action(self.request.user, f"deleted file {file_name}", ip_address=get_client_ip(self.request))

    @action(detail=True, methods=["patch"], url_path="status", parser_classes=[JSONParser])
    def update_status(self, request, pk=None):
        file = self.get_object()
        previous_status = file.status

        serializer = FileStatusSerializer(file, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        new_status = serializer.data.get("status")

        if new_status == "rejected" and previous_status != "rejected":
            user = file.owner
            if user.phone_number:
                try:
                    send_rejection_sms(user.phone_number, file.file.name, user=user)
                except Exception as e:
                    log_action(request.user, f"failed to send rejection SMS to {user.username}: {str(e)}", status="error")

        log_action(request.user, f"updated status of file {file.file.name} to {new_status}", ip_address=get_client_ip(request))
        return Response(serializer.data)
    
    @action(detail=True, methods=["get"], url_path="content")
    def get_content(self, request, pk=None):
        file_obj = self.get_object()
        print(f"User: {request.user}, Role: {request.user.role}, File: {file_obj.file.name}")

        if request.user.role not in ["admin", "viewer" , "client"]:
            return Response({"detail": "Forbidden"}, status=403)

        if file_obj.parsed_content:
            return Response({"pages": file_obj.parsed_content})

        file_name = file_obj.file.name.lower()
        try:
            # --- CSV ---
            if file_name.endswith(".csv"):
                file_obj.file.seek(0) 
                file_data = file_obj.file.read()
                decoded_data = file_data.decode("utf-8").splitlines()
                reader = csv.reader(decoded_data)
                return Response({"pages": [{"page_number": 1, "content": list(reader)}]})

            # --- XLSX ---
            elif file_name.endswith(".xlsx"):
                file_obj.file.seek(0)
                file_bytes = io.BytesIO(file_obj.file.read())
                wb = load_workbook(file_bytes, read_only=True)
                ws = wb.active
                rows = [[str(cell) if cell is not None else "" for cell in row] for row in ws.iter_rows(values_only=True)]
                return Response({"pages": [{"page_number": 1, "content": rows}]})

            # --- PDF ---
            elif file_name.endswith(".pdf"):
                import pdfplumber
                pages_data = []

                def is_number(val):
                    try:
                        float(val)
                        return True
                    except ValueError:
                        return False

                file_obj.file.seek(0)
                with pdfplumber.open(file_obj.file) as pdf:
                    for i, page in enumerate(pdf.pages, start=1):
                        page_data = {"page_number": i}

                        text = page.extract_text()
                        if text:
                            page_data["text"] = text

                        lines = text.splitlines() if text else []
                        structured_table = {"main_headers": [], "sub_headers": [], "rows": []}

                        if lines:
                            header_idx = None
                            for idx, line in enumerate(lines):
                                if "Emp." in line and "DUTY" in line:
                                    header_idx = idx
                                    break

                            if header_idx is not None:
                                structured_table["main_headers"] = [
                                    "Emp. No",
                                    "Name",
                                    "Duty (By Days)",
                                    "Late",
                                    "UT",
                                    "Work (By Hrs)",
                                    "Day-Off (By Hours)",
                                    "SH (By Hrs)",
                                    "LH (By Hrs)",
                                    "Day-Off - SH (By Hrs)",
                                    "Day-Off - LH (By Hrs)"
                                ]
                                structured_table["sub_headers"] = [
                                    [""], [""],
                                    ["WRK", "ABS", "LV", "HOL", "RES"],
                                    [""], [""],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                ]

                                data_lines = lines[header_idx + 1:]
                                expected_cols = sum(len(group) for group in structured_table["sub_headers"])

                                for dl in data_lines:
                                    parts = dl.split()
                                    if not parts or not parts[0].isdigit():
                                        continue

                                    emp_no = parts[0]
                                    name_parts, numbers = [], []
                                    found_number = False

                                    for p in parts[1:]:
                                        if is_number(p):
                                            found_number = True
                                            numbers.append(f"{float(p):.2f}")
                                        elif not found_number:
                                            name_parts.append(p)

                                    clean_name = " ".join(name_parts).strip()
                                    padded_numbers = numbers + ["0.00"] * (expected_cols - len(numbers))
                                    row = [emp_no, clean_name] + padded_numbers[:expected_cols]
                                    structured_table["rows"].append(row)

                        if structured_table["rows"]:
                            page_data["tables"] = [structured_table]

                        pages_data.append(page_data)

                return Response({"pages": pages_data})

            # --- Images ---
            elif file_name.endswith((".jpg", ".jpeg", ".png")):
                import cv2, numpy as np, easyocr
                file_obj.file.seek(0)
                file_bytes = np.asarray(bytearray(file_obj.file.read()), dtype=np.uint8)
                img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
                if img is None:
                    raise ValueError("Failed to decode image")

                reader = easyocr.Reader(["en"])
                ocr_result = reader.readtext(img)

                rows_dict = {}
                row_threshold = 10
                for (bbox, text, conf) in ocr_result:
                    top = int(bbox[0][1])
                    left = int(bbox[0][0])
                    if not text.strip():
                        continue
                    found = False
                    for key in rows_dict:
                        if abs(key - top) < row_threshold:
                            rows_dict[key].append((left, text))
                            found = True
                            break
                    if not found:
                        rows_dict[top] = [(left, text)]

                table = []
                for top in sorted(rows_dict.keys()):
                    row_words = sorted(rows_dict[top], key=lambda x: x[0])
                    table.append([w[1] for w in row_words])

                return Response({"pages": [{"page_number": 1, "content": table}]})

            else:
                return Response({"detail": "Unsupported file type"}, status=400)

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"detail": f"Failed to read file: {str(e)}"}, status=400)

    @action(detail=True, methods=["patch"], url_path="update-content")
    def update_content(self, request, pk=None):
        file_obj = self.get_object()
        if request.user.role not in ["admin", "viewer"]:
            return Response({"detail": "Forbidden"}, status=403)

        pages = request.data.get("pages")
        content = request.data.get("content") 

        if not pages and not content:
            return Response({"detail": "No content provided"}, status=400)

        if pages and not content:
            content = []
            for page in pages:
                if page.get("tables"):
                    for table in page["tables"]:
                        rows = table.get("rows", [])
                        sub_headers = table.get("sub_headers") or []
                        expected_cols = sum(len(group) if isinstance(group, list) else 1 for group in sub_headers)

                        for row in rows:
                            normalized = []
                            for c in range(expected_cols):
                                if c < len(row):
                                    val = row[c]
                                    try:
                                        normalized.append(float(val))
                                    except (ValueError, TypeError):
                                        normalized.append(str(val or "0.00"))
                                else:
                                    normalized.append("0.00") 
                            content.append(normalized)
                elif page.get("content"): 
                    content.extend(page["content"])

        file_name = file_obj.file.name.lower()

        def format_numeric(val):
            if val is None or val == "":
                return "0.00"
            try:
                num = float(val)
                return f"{num:.2f}"
            except (ValueError, TypeError):
                return str(val)

        try:
            # --- CSV ---
            if file_name.endswith(".csv"):
                from django.core.files.base import ContentFile
                import csv, io
                output = io.StringIO()
                writer = csv.writer(output)
                for row in content:
                    writer.writerow([format_numeric(cell) if isinstance(cell, (int, float, str)) else str(cell) for cell in row])
                file_obj.file.save(file_obj.file.name, ContentFile(output.getvalue()), save=True)

            # --- XLSX ---
            elif file_name.endswith(".xlsx"):
                from openpyxl import Workbook
                from django.core.files.base import ContentFile
                from io import BytesIO

                wb = Workbook()
                ws = wb.active
                for row in content:
                    new_row = []
                    for cell in row:
                        try:
                            new_row.append(float(cell))
                        except (ValueError, TypeError):
                            new_row.append(str(cell or "0.00"))
                    ws.append(new_row)
                stream = BytesIO()
                wb.save(stream)
                file_obj.file.save(file_obj.file.name, ContentFile(stream.getvalue()), save=True)

            # --- PDF ---
            elif file_name.endswith(".pdf"):
                from reportlab.lib.pagesizes import letter
                from reportlab.pdfgen import canvas
                from reportlab.pdfbase.pdfmetrics import stringWidth
                from reportlab.lib.colors import lightgrey, black
                from django.core.files.base import ContentFile
                from io import BytesIO

                buffer = BytesIO()
                p = canvas.Canvas(buffer, pagesize=letter)

                for page in pages:
                    y_offset = 750
                    page_has_content = False

                    if "text" in page and page["text"]:
                        p.setFont("Helvetica", 10)
                        for line in page["text"].splitlines():
                            p.drawString(50, y_offset, line)
                            y_offset -= 15
                            page_has_content = True
                            if y_offset < 50:
                                p.showPage()
                                y_offset = 750
                        p.showPage()

                    if "tables" in page and page["tables"]:
                        for table in page["tables"]:
                            main_headers = table.get("main_headers", [])
                            sub_headers = table.get("sub_headers", [])
                            rows = table.get("rows", [])

                            flat_sub_headers = []
                            for group in sub_headers:
                                flat_sub_headers.extend(group if isinstance(group, list) else [group])

                            all_rows = [main_headers] + [flat_sub_headers] + rows
                            if not all_rows:
                                continue

                            num_cols = max(len(r) for r in all_rows)
                            col_widths = []
                            row_height = 20

                            for col_idx in range(num_cols):
                                max_width = max(
                                    stringWidth(format_numeric(row[col_idx]) if col_idx < len(row) else "0.00", "Helvetica", 8)
                                    for row in all_rows
                                )
                                col_widths.append(max_width + 20)

                            x_offset, y_start = 50, y_offset

                            # Main headers
                            x = x_offset
                            for j, h in enumerate(main_headers):
                                span = len(sub_headers[j]) if j < len(sub_headers) else 1
                                span_width = sum(col_widths[j:j+span])
                                p.setFillColor(lightgrey)
                                p.rect(x, y_start, span_width, -row_height, fill=1, stroke=1)
                                p.setFillColor(black)
                                p.setFont("Helvetica-Bold", 8)
                                p.drawCentredString(x + span_width / 2, y_start - row_height + 15, h)
                                x += span_width
                            y_start -= row_height

                            # Sub headers
                            x = x_offset
                            for group in sub_headers:
                                for sh in (group if isinstance(group, list) else [group]):
                                    width = col_widths[sub_headers.index(group)]
                                    p.rect(x, y_start, width, -row_height, fill=0, stroke=1)
                                    p.setFont("Helvetica-Bold", 8)
                                    p.drawCentredString(x + width / 2, y_start - row_height + 15, sh)
                                    x += width
                            y_start -= row_height

                            # Table rows
                            for row in rows:
                                if y_start < 50:
                                    p.showPage()
                                    y_start = 750
                                x = x_offset
                                for j in range(num_cols):
                                    text = format_numeric(row[j]) if j < len(row) else "0.00"
                                    p.rect(x, y_start, col_widths[j], -row_height, fill=0, stroke=1)
                                    p.setFont("Helvetica", 8)
                                    p.drawString(x + 2, y_start - row_height + 15, text)
                                    x += col_widths[j]
                                y_start -= row_height

                            y_offset = y_start
                        p.showPage()

                    if not page_has_content:
                        p.setFont("Helvetica", 10)
                        p.drawString(50, 750, "No content available")
                        p.showPage()

                p.save()
                buffer.seek(0)
                file_obj.file.save(file_obj.file.name, ContentFile(buffer.read()), save=True)

            # --- Images ---
            elif file_name.endswith((".jpg", ".jpeg", ".png")):
                import cv2
                import numpy as np
                from django.core.files.base import ContentFile

                rows = content
                cell_width, cell_height = 200, 50
                font, font_scale, thickness = cv2.FONT_HERSHEY_SIMPLEX, 0.7, 1

                n_rows = len(rows)
                n_cols = max(len(r) for r in rows) if rows else 0
                if n_rows == 0 or n_cols == 0:
                    return Response({"detail": "No content to update"}, status=400)

                img_height = n_rows * cell_height + 2
                img_width = n_cols * cell_width + 2
                img = np.ones((img_height, img_width, 3), dtype=np.uint8) * 255

                for i, row in enumerate(rows):
                    for j, cell in enumerate(row):
                        x1, y1 = j * cell_width, i * cell_height
                        x2, y2 = x1 + cell_width, y1 + cell_height
                        cv2.rectangle(img, (x1, y1), (x2, y2), (0, 0, 0), 1)
                        text = str(cell)
                        (tw, th), _ = cv2.getTextSize(text, font, font_scale, thickness)
                        text_x = x1 + (cell_width - tw) // 2
                        text_y = y1 + (cell_height + th) // 2
                        cv2.putText(img, text, (text_x, text_y), font, font_scale, (0, 0, 0), thickness)

                _, buffer = cv2.imencode(".png", img)
                file_obj.file.save(file_obj.file.name, ContentFile(buffer.tobytes()), save=True)

            else:
                return Response({"detail": "Unsupported file type"}, status=400)

            file_obj.parsed_content = pages or content
            file_obj.save(update_fields=["parsed_content"])
            return Response({"detail": "Content updated successfully"})

        except Exception as e:
            return Response({"detail": f"Failed to save content: {str(e)}"}, status=400)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    files_pending = File.objects.filter(status="pending").count()
    files_rejected = File.objects.filter(status="rejected").count()
    files_approved = File.objects.filter(status="verified").count()
    active_users = User.objects.filter(is_active=True).count()
    
    return Response({
        "filesPending": files_pending,
        "filesApproved": files_approved,
        "filesRejected": files_rejected,
        "activeUsers": active_users
    })

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_files_report(request):
    format = request.GET.get("format", "csv")
    files = File.objects.all().order_by("-uploaded_at")
    
    if format == "csv":
        response = HttpResponse(content_type="text/csv")
        response['Content-Disposition'] = 'attachment; filename="files_report.csv"'
        writer = csv.writer(response)
        writer.writerow(["ID", "Filename", "Owner", "Status", "Uploaded At"])
        for f in files:
            writer.writerow([f.id, f.file.name, f.owner.username, f.status, f.uploaded_at])
        return response
    
    elif format == "pdf":
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = 'attachment; filename="files_report.pdf"'
        p = canvas.Canvas(response)
        y = 800
        p.drawString(50, y, "Files Report")
        y -= 25
        for f in files:
            p.drawString(50, y, f"{f.id} | {f.file.name} | {f.owner.username} | {f.status} | {f.uploaded_at}")
            y -= 20
        p.showPage()
        p.save()
        return response
    
    else:
        return Response({"detail": "Unsupported format"}, status=400)
    
class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all().order_by("-timestamp")
    serializer_class = AuditLogSerializer
    permission_classes = [IsAdmin]  

class SystemSettingsViewSet(viewsets.ModelViewSet):
    queryset = SystemSettings.objects.all()
    serializer_class = SystemSettingsSerializer
    permission_classes = [IsAdmin] 

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def file_stats(request):
    period = request.query_params.get("period", "day")  

    if period == "month":
        trunc = TruncMonth("uploaded_at")
    elif period == "week":
        trunc = TruncWeek("uploaded_at")
    else:
        trunc = TruncDay("uploaded_at")

    stats = (
        File.objects
        .annotate(period=trunc)
        .values("period")
        .annotate(
            pending=Count("id", filter=Q(status="pending")),
            verified=Count("id", filter=Q(status="verified")),
            rejected=Count("id", filter=Q(status="rejected")),
        )
        .order_by("period")
    )

    return Response(stats)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def file_audit_logs(request):
    logs = AuditLog.objects.filter(
        action__in=["upload", "delete", "update"]
    ).order_by("-timestamp")

    data = [
        {
            "user": log.user.username if log.user else "Unknown",
            "role": log.user.role if log.user else "Unknown",
            "action": log.action,
            "status": log.status,
            "ip_address": log.ip_address,
            "timestamp": log.timestamp,
        }
        for log in logs
    ]
    return Response(data)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def rejected_files(request):
    user = request.user
    files = File.objects.filter(owner=user, status="rejected").order_by("-uploaded_at")
    serializer = FileSerializer(files, many=True)
    return Response(serializer.data)

def clean_value(value):
    """Convert NaN or missing values to None for JSON storage."""
    if pd.isna(value) or (isinstance(value, float) and math.isnan(value)):
        return None
    return value

@api_view(['POST'])
@permission_classes([IsAdminUser])
def upload_employee_excel(request):
    file = request.FILES.get('file')
    if not file:
        return Response({"detail": "No file uploaded."}, status=400)
    
    try:
        df = pd.read_excel(file, header=0, dtype=str)
        df.columns = [col.strip() for col in df.columns]  
        df_cleaned = df.applymap(clean_value)

        added_count = 0
        updated_count = 0

        def to_float(val):
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        for _, row in df_cleaned.iterrows():
            employee_code_raw = row.get('Employee Code')
            employee_code = str(employee_code_raw).strip().zfill(5) if employee_code_raw else ""

            data = {
                'employee_code': employee_code,
                'employee_name': str(row.get('EmployeeName', '')).strip(),
                'total_hours': to_float(row.get('Total Hours')),
                'nd_reg_hrs': to_float(row.get('ND Reg Hrs')),
                'absences': to_float(row.get('Absences')),
                'tardiness': to_float(row.get('Tardiness')),
                'undertime': to_float(row.get('Undertime')),
                'ot_regular': to_float(row.get('OTRegular')),
                'nd_ot_reg': to_float(row.get('ND OT Reg')),
                'ot_restday': to_float(row.get('OT Restday')),
                'nd_restday': to_float(row.get('ND Restday')),
                'ot_rest_excess': to_float(row.get('OT RestExcess')),
                'nd_rest_excess': to_float(row.get('ND Restday Excess')),
                'ot_special_hday': to_float(row.get('OTSpecialHday')),
                'nd_special_hday': to_float(row.get('ND SpecialHday')),
                'ot_shday_excess': to_float(row.get('OT SHdayExcess')),
                'nd_shday_excess': to_float(row.get('ND SHday Excess')),
                'ot_legal_holiday': to_float(row.get('OT LegalHoliday')),
                'special_holiday': to_float(row.get('Special Holiday')),
                'ot_leghol_excess': to_float(row.get('OTLegHol Excess')),
                'nd_leghol_excess': to_float(row.get('ND LegHol Excess')),
                'ot_sh_on_rest': to_float(row.get('OT SHday on Rest')),
                'nd_sh_on_rest': to_float(row.get('ND SH on Rest')),
                'ot_sh_on_rest_excess': to_float(row.get('OT SH on Rest Excess')),
                'nd_sh_on_rest_excess': to_float(row.get('ND SH on Rest Excess')),
                'leg_h_on_rest_day': to_float(row.get('LegH on Rest Day')),
                'nd_leg_h_on_restday': to_float(row.get('ND LegH on Restday')),
                'ot_leg_h_on_rest_excess': to_float(row.get('OT LegH on Rest Excess')),
                'nd_leg_h_on_rest_excess': to_float(row.get('ND LegH on Rest Excess')),
                'vacleave_applied': to_float(row.get('VacLeave_Applied')),
                'sickleave_applied': to_float(row.get('SickLeave_Applied')),
                'back_pay_vl': to_float(row.get('Back Pay VL')),
                'back_pay_sl': to_float(row.get('Back Pay SL')),
                'ot_regular_excess': to_float(row.get('OTRegular Excess')),
                'nd_ot_reg_excess': to_float(row.get('ND OT Reg Excess')),
                'legal_holiday': to_float(row.get('Legal Holiday')),
                'nd_legal_holiday': to_float(row.get('ND Legal Holiday')),
                'overnight_rate': to_float(row.get('Overnight Rate')),
                'project': str(row.get('PROJECT', '')).strip(),
            }

            if employee_code:
                obj, created = EmployeeDirectory.objects.update_or_create(
                    employee_code=employee_code,
                    defaults=data
                )
                if created:
                    added_count += 1
                else:
                    updated_count += 1
            else:
                obj = EmployeeDirectory.objects.filter(employee_name=data['employee_name'], employee_code="").first()
                if obj:
                    for key, value in data.items():
                        setattr(obj, key, value)
                    obj.save()
                    updated_count += 1
                else:
                    EmployeeDirectory.objects.create(**data)
                    added_count += 1

        return Response({
            "detail": f"{added_count} new employees added, {updated_count} employees updated."
        })

    except Exception as e:
        return Response({"detail": str(e)}, status=400)
    
@api_view(['GET'])
@permission_classes([IsAdminUser])
def list_employees(request):
    employees = EmployeeDirectory.objects.all().order_by("id")
    serializer = EmployeeDirectorySerializer(employees, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAdminUser])
def add_employee(request):
    employee_code = request.data.get("employee_code")
    employee_name = request.data.get("employee_name")

    if not employee_code or not employee_name:
        return Response({"detail": "Employee code and name are required."}, status=400)

    employee_code = str(employee_code).strip().zfill(5)
    employee_name = employee_name.strip()

    obj, created = EmployeeDirectory.objects.get_or_create(
        employee_code=employee_code,
        defaults={"employee_name": employee_name}
    )

    if not created:
        return Response({"detail": "Employee already exists."}, status=400)

    return Response({"detail": "Employee added successfully!"})

@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def delete_employee(request, employee_code):
    employee_code_str = str(employee_code)
    padded_code = employee_code_str.zfill(5)

    try:
        employee = EmployeeDirectory.objects.get(
            Q(employee_code=padded_code) | Q(employee_code=employee_code_str)
        )
        employee.delete()
        return Response({"detail": "Employee deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    except EmployeeDirectory.DoesNotExist:
        return Response({"detail": "Employee not found."}, status=status.HTTP_404_NOT_FOUND)

@api_view(['PUT'])
@permission_classes([IsAdminUser])
def update_employee(request, employee_code):
    employee_code_str = str(employee_code)
    padded_code = employee_code_str.zfill(5)
    
    try:
        employee = EmployeeDirectory.objects.get(
            Q(employee_code=padded_code) | Q(employee_code=employee_code_str)
        )
    except EmployeeDirectory.DoesNotExist:
        return Response({"detail": "Employee not found"}, status=404)

    numeric_fields = [
        "total_hours", "nd_reg_hrs", "absences", "tardiness", "undertime",
        "ot_regular", "nd_ot_reg", "ot_restday", "nd_restday", "ot_rest_excess",
        "nd_rest_excess", "ot_special_hday", "nd_special_hday", "ot_shday_excess",
        "nd_shday_excess", "ot_legal_holiday", "special_holiday", "ot_leghol_excess",
        "nd_leghol_excess", "ot_sh_on_rest", "nd_sh_on_rest", "ot_sh_on_rest_excess",
        "nd_sh_on_rest_excess", "leg_h_on_rest_day", "nd_leg_h_on_restday",
        "ot_leg_h_on_rest_excess", "nd_leg_h_on_rest_excess", "vacleave_applied",
        "sickleave_applied", "back_pay_vl", "back_pay_sl", "ot_regular_excess",
        "nd_ot_reg_excess", "legal_holiday", "nd_legal_holiday", "overnight_rate",
    ]

    for field, value in request.data.items():
        if hasattr(employee, field):
            if field in numeric_fields:
                if value in ["", None]:
                    setattr(employee, field, None)
                else:
                    try:
                        setattr(employee, field, Decimal(value))
                    except InvalidOperation:
                        setattr(employee, field, None)
            else:
                setattr(employee, field, value)

    employee.save()
    return Response({"detail": "Employee updated successfully"})

def safe_number(val, default=0):
    """Convert pandas/Excel values into a safe float or default."""
    if pd.isna(val):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def safe_string(val):
    """Convert to string if not NaN, else return None."""
    if pd.isna(val):
        return None
    return str(val).strip()

class DTRFileViewSet(viewsets.ModelViewSet):
    queryset = DTRFile.objects.all().order_by("-uploaded_at")
    serializer_class = DTRFileSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)

    @action(detail=True, methods=["post"])
    def parse(self, request, pk=None):
        dtr_file = self.get_object()
        file_path = dtr_file.file.path
        dtr_file.entries.all().delete() 

        try:
            df = pd.read_excel(file_path, header=None)

            start_date_val = df.iat[8, 4] if not pd.isna(df.iat[8, 4]) else None
            end_date_val = df.iat[9, 4] if not pd.isna(df.iat[9, 4]) else None

            dtr_file.start_date = pd.to_datetime(start_date_val).date() if start_date_val else None
            dtr_file.end_date = pd.to_datetime(end_date_val).date() if end_date_val else None
            dtr_file.save()

            employee_df = df.iloc[13:, :]

            for _, row in employee_df.iterrows():
                name = row[2]
                emp_no = row[4]

                if pd.isna(name) or pd.isna(emp_no):
                    continue
                try:
                    code = str(int(emp_no)).zfill(5)
                except (ValueError, TypeError):
                    continue  

                daily_data = {}
                if dtr_file.start_date:
                    for idx, col in enumerate(range(8, 24)):
                        day = dtr_file.start_date + timedelta(days=idx)
                        val = row[col]
                        daily_data[str(day)] = None if pd.isna(val) else val

                DTREntry.objects.create(
                    dtr_file=dtr_file,
                    full_name=safe_string(name),
                    employee_no=code,
                    area=safe_string(row[5]),
                    daily_data=daily_data,
                    total_days=safe_number(row[24]),
                    total_hours=safe_number(row[25]),
                    undertime_minutes=safe_number(row[26]),
                    regular_ot=safe_number(row[27]),
                    legal_holiday=safe_number(row[28]),
                    unworked_reg_holiday=safe_number(row[29]),
                    special_holiday=safe_number(row[30]),
                    night_diff=safe_number(row[31]),
                )

            return Response({"message": "DTR file parsed successfully."})

        except Exception as e:
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["get"])
    def content(self, request, pk=None):
        dtr_file = self.get_object()
        entries = dtr_file.entries.all()
        serializer = DTREntrySerializer(entries, many=True)
        return Response({
            "start_date": dtr_file.start_date,
            "end_date": dtr_file.end_date,
            "rows": serializer.data
        })
    
    @action(detail=False, methods=["post"], url_path="sync-all")
    def sync_all_files(self, request):
        """
        Sync all DTR entries to EmployeeDirectory.
        Optional request data:
            start_date, end_date -> filter which DTR files to consider
        """
        start_date_str = request.data.get("start_date")
        end_date_str = request.data.get("end_date")

        start_date = pd.to_datetime(start_date_str).date() if start_date_str else None
        end_date = pd.to_datetime(end_date_str).date() if end_date_str else None

        dtr_files = DTRFile.objects.all().order_by("start_date")
        if start_date:
            dtr_files = dtr_files.filter(end_date__gte=start_date)
        if end_date:
            dtr_files = dtr_files.filter(start_date__lte=end_date)

        aggregated = {}

        for dtr_file in dtr_files:
            for entry in dtr_file.entries.all():
                code = str(entry.employee_no).zfill(5) if entry.employee_no else None
                name = entry.full_name.strip() if entry.full_name else None
                if not code or not name:
                    continue

                if code not in aggregated:
                    aggregated[code] = {
                        "employee_name": name,
                        "total_hours": 0,
                        "undertime": 0,
                        "ot_regular": 0,
                        "legal_holiday": 0,
                        "special_holiday": 0,
                        "nd_reg_hrs": 0,
                        "date_covered_start": dtr_file.start_date,
                        "date_covered_end": dtr_file.end_date,
                    }

                aggregated[code]["total_hours"] += entry.total_hours or 0
                aggregated[code]["undertime"] += entry.undertime_minutes or 0
                aggregated[code]["ot_regular"] += entry.regular_ot or 0
                aggregated[code]["legal_holiday"] += entry.legal_holiday or 0
                aggregated[code]["special_holiday"] += entry.special_holiday or 0
                aggregated[code]["nd_reg_hrs"] += entry.night_diff or 0

                if dtr_file.start_date and (
                    aggregated[code]["date_covered_start"] is None
                    or dtr_file.start_date < aggregated[code]["date_covered_start"]
                ):
                    aggregated[code]["date_covered_start"] = dtr_file.start_date

                if dtr_file.end_date and (
                    aggregated[code]["date_covered_end"] is None
                    or dtr_file.end_date > aggregated[code]["date_covered_end"]
                ):
                    aggregated[code]["date_covered_end"] = dtr_file.end_date

        created, updated = 0, 0
        for code, data in aggregated.items():
            date_covered = None
            if data.get("date_covered_start") and data.get("date_covered_end"):
                start_str = data["date_covered_start"].strftime("%b %d, %Y")
                end_str = data["date_covered_end"].strftime("%b %d, %Y")
                date_covered = f"{start_str} → {end_str}"

            emp, is_created = EmployeeDirectory.objects.update_or_create(
                employee_code=code,
                defaults={
                    **{k: v for k, v in data.items() if k not in ["date_covered_start", "date_covered_end"]},
                    "date_covered": date_covered,
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1

        return Response({
            "detail": f"All DTR files synced: {created} new, {updated} updated."
        })

class DTREntryViewSet(viewsets.ModelViewSet):
    queryset = DTREntry.objects.all().order_by("full_name")
    serializer_class = DTREntrySerializer
    permission_classes = [IsAuthenticated]
